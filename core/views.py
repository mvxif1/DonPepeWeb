from datetime import date
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from .models import Direccion, Estado, ItemCarrito, ProductoCarrito, TipoDespacho, Usuario, Producto, Venta, Categoria, DetalleVenta, Rol, Region, Comuna, Review
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Sum
from django.shortcuts import render
from .forms import ReviewForm
# Create your views here.

def inicio(request):
    carrito = []
    total = 0
    carrito_count = 0
    
    if request.user.is_authenticated:
        carrito = ItemCarrito.objects.filter(usuario=request.user)
        total = sum(item.producto.precio * item.cantidad for item in carrito)
        carrito_count = sum(item.cantidad for item in carrito)
    reviews = Review.objects.all().order_by('-date_posted')
    contexto = {
        'carrito': carrito,
        'total': total,
        'carrito_count': carrito_count,
        'reviews': reviews
    }
    
    return render(request, 'core/inicio.html', contexto)

def register(request):
    if request.user.is_authenticated:
        return redirect('inicio')
    
    if request.method == "POST":
        nombreUser = request.POST['nombre']
        apellidoUser = request.POST['apellido']
        rutUser = request.POST['rut']
        telefonoUser = request.POST['telefono']
        emailUser = request.POST['email']
        claveUser = request.POST['password']
        confclaveUser = request.POST['confirmPassword']
        
        # Validaciones
        if User.objects.filter(username=emailUser).exists():
            messages.error(request, "Este correo ya está registrado!")
            return redirect('register')
        
        if Usuario.objects.filter(rut=rutUser).exists():
            messages.error(request, "Este RUT ya está registrado!")
            return redirect('register')
        
        if claveUser != confclaveUser:
            messages.error(request, "La contraseña no coincide")
            return redirect('register')
                
        rol = Rol.objects.get(idRol=2)  # Suponiendo que el rol "2" es el rol que deseas asignar
        usuario = Usuario.objects.create(rut=rutUser, nombre=nombreUser, apellido=apellidoUser, telefono=telefonoUser, correo=emailUser, clave=confclaveUser, rol=rol)

        user = User.objects.create_user(username = emailUser, email=emailUser, password= claveUser)
        user.first_name = nombreUser
        user.last_name = apellidoUser
        user.is_staff = True
        user.save()
        messages.success(request, 'Cuenta creada con éxito.')
        
        return redirect('login_user')

    else:
        return render(request, 'core/register.html')
def login_user(request):
    if request.user.is_authenticated:
        return redirect('inicio')
    
    if request.method == 'POST':
        email = request.POST['username']
        password = request.POST['pass1']

        user = authenticate(request, username=email, password=password)
        print("Usuario autenticado: ", user)
        if user is not None:
            login(request, user)
            if user.is_superuser:  # Verificar si el usuario es un superusuario
                # Asignar el rol de administrador al superusuario
                rol_administrador = Rol.objects.get(idRol=1)  # Suponiendo que el id del rol de administrador es 1
                user.rol = rol_administrador
            else:
                # Asignar el rol de usuario al usuario normal
                rol_usuario = Rol.objects.get(idRol=2)  # Suponiendo que el id del rol de usuario es 2
                user.rol = rol_usuario
            user.save()
            messages.success(request, 'Inicio de sesión exitoso.')
            return redirect('inicio')
        else:
            messages.error(request, 'Correo o contraseña incorrectos.')

    return render(request, 'core/login_user.html')

def cerrarsesion(request):
    logout(request)
    messages.success(request, "Sesión cerrada correctamente!!")
    return redirect('inicio')


@login_required
def editarperfil(request):
    user = request.user
    usuario = Usuario.objects.get(correo=user.username)
    context = {
        'usuario': usuario
    }
    return render(request, 'core/editarperfil.html', context)

@login_required
def actualizarperfil(request):
    user = request.user
    usuario = Usuario.objects.get(correo=user.username)
    if request.method == "POST":
        nombreUsuario = request.POST['nombre']
        apellidoUsuario = request.POST['apellido']
        telefonoUsuario = request.POST['telefono']
        emailUsuario = request.POST['email']
        
        # Verificar si el correo electrónico ha sido cambiado
        if emailUsuario != user.username:
            # Si el nuevo correo electrónico es diferente al actual del usuario, realizar validación
            if User.objects.filter(username=emailUsuario).exists():
                messages.error(request, "Este correo ya está registrado!")
                return redirect('editarperfil')
        
        usuario.nombre = nombreUsuario
        usuario.apellido = apellidoUsuario
        usuario.telefono = telefonoUsuario
        usuario.correo = emailUsuario
        user.username = emailUsuario
        user.email= emailUsuario
        
        usuario.save()
        user.save()

        messages.success(request, 'Cuenta actualizada con éxito.')
        return redirect('editarperfil')
    
    context = {
        'usuario': usuario
    }
    return render(request, 'core/editarperfil.html', context)


def administrador(request):
    return render(request, 'core/administrador.html')

def agregar(request):
    categoriaProducto = Categoria.objects.all()
    contexto = {
        "categorias" : categoriaProducto
    }

    return render(request, 'core/agregar.html', contexto)

def ingresarproducto(request):

    idProducto = request.POST['id']
    nombreProducto = request.POST['nombre']
    stockProducto = request.POST['stock']
    descripcion = request.POST['descripcion']
    foto = request.FILES['foto']
    precio = request.POST['precio']
    categoria = request.POST['categoria']

    categoriaP= Categoria.objects.get(idCategoria= categoria)

    producto = Producto.objects.create(codProducto= idProducto, nombreP= nombreProducto, stock= stockProducto, descipcion= descripcion, foto= foto, precio= precio, categoria= categoriaP)
    messages.success(request, 'Producto ingresado correctamente.')
    return redirect('agregar')

def listaproducto(request):
    productoListado = Producto.objects.all()
    categorias = Categoria.objects.all()

    # Anotamos la cantidad total vendida para cada producto
    productos_vendidos = DetalleVenta.objects.values('producto_id').annotate(total_vendido=Sum('cantidad')).order_by('-total_vendido')

    # Agregamos la cantidad vendida a cada producto
    for producto in productoListado:
        producto.cantidad_vendida = 0
        for pv in productos_vendidos:
            if pv['producto_id'] == producto.codProducto:
                producto.cantidad_vendida = pv['total_vendido']
                break

    # Filtramos los productos con stock menor a 15
    productos_bajo_stock = productoListado.filter(stock__lt=15)
    
    contexto = {
        "categorias": categorias,
        "productos": productoListado,
        'productos_bajo_stock': productos_bajo_stock
    }   
    return render(request, 'core/listaproducto.html', contexto)

def exportar_productos_excel(request):
    productos = Producto.objects.all()
    
    # Crear un libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilo de encabezado
    header_font = Font(bold=True, size=12)
    alignment_center = Alignment(horizontal="center", vertical="center")
    
    # Definir estilo de borde delgado
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Definir colores de fondo para alternar en los títulos
    fill_gray_light = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    fill_gray_dark = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    
    # Desplazar la tabla para que comience en B2
    start_col = 2  # Comenzar en la columna B
    start_row = 2  # Comenzar en la fila 2
    
    # Definir encabezados
    headers = ["ID Producto", "Nombre", "Stock", "Precio", "Categoría"]
    for col_num, header in enumerate(headers, start_col):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border  # Aplicar borde al encabezado
        # Alternar colores
        cell.fill = fill_gray_light if col_num % 2 == 0 else fill_gray_dark
    
    # Añadir datos de productos
    for row_num, producto in enumerate(productos, start_row + 1):
        cells = [
            ws.cell(row=row_num, column=start_col, value=producto.codProducto),
            ws.cell(row=row_num, column=start_col + 1, value=producto.nombreP),
            ws.cell(row=row_num, column=start_col + 2, value=producto.stock),
            ws.cell(row=row_num, column=start_col + 3, value=producto.precio),
            ws.cell(row=row_num, column=start_col + 4, value=producto.categoria.nombreCa if hasattr(producto.categoria, 'nombreCa') else '')
        ]
        for cell in cells:
            cell.alignment = alignment_center
            cell.border = thin_border  # Aplicar borde a las celdas de datos
            cell.font = Font(size=12)  # Establecer tamaño de fuente en 12
    
    # Ajustar el ancho de las columnas basado en el contenido
    for col in ws.iter_cols(min_col=start_col, max_col=start_col + len(headers) - 1):
        max_length = 0
        column = col[0].column_letter  # Obtener el nombre de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajuste adicional del ancho
        ws.column_dimensions[column].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

def buscar_productos(request):
    query = request.GET.get('q')
    if query:
        productos = Producto.objects.filter(nombreP__icontains=query)
    else:
        productos = Producto.objects.all()
    
    context = {
        'productos': productos
    }
    return render(request, 'core/listaproducto.html', context)

def editarproducto(request, id_producto):
    producto = Producto.objects.get(codProducto = id_producto)
    listaCategoria = Categoria.objects.all()
    contexto = {
        "producto" : producto,
        "listacategoria" : listaCategoria
    }
    return render(request, 'core/editarproducto.html', contexto)

def actualizaproducto(request):
    if request.method == "POST":
        idProducto = request.POST['id']
        nombreProducto = request.POST['nombre']
        stockProducto = request.POST['stock']
        descripcion = request.POST['descripcion']
        precio = request.POST['precio']
        categoria = request.POST['categoria']

        producto = Producto.objects.get(codProducto=idProducto)
        categoriaP = Categoria.objects.get(idCategoria=categoria)

        producto.nombreP = nombreProducto
        producto.stock = stockProducto
        producto.descipcion = descripcion
        producto.precio = precio
        producto.categoria = categoriaP

        if 'imagen' in request.FILES:
            producto.foto = request.FILES['imagen']

        producto.save()
        messages.success(request, 'Producto actualizado correctamente.')
        return redirect('listaproducto')


def borrarproducto(request, id_producto):
    productoborrar = Producto.objects.get(codProducto = id_producto)
    productoborrar.delete()
    messages.success(request, 'Producto eliminado correctamente.')
    return redirect('listaproducto')


def listausuarios(request):
    usuariosListado = Usuario.objects.exclude(rut=None).exclude(rut='').all()
    contexto = {
        "usuarios": usuariosListado
    }
    return render(request, 'core/listausuarios.html', contexto)

def borrarperfil(request, correo):
    try:
        # Eliminar el usuario de Django por su username (en este caso, asumiendo que el username es igual al correo)
        user = User.objects.get(username=correo)
        user.delete()

        # Eliminar el usuario de la tabla Usuario
        perfilborrar = Usuario.objects.get(correo=correo)
        perfilborrar.delete()

        messages.success(request, 'Perfil eliminado correctamente.')
    except User.DoesNotExist:
        messages.error(request, 'El usuario no existe en la base de datos.')

    return redirect('listausuarios')

@login_required
def productos(request):
    nombre = request.GET.get('nombre', '')
    categoria_id = request.GET.get('categoria', '')
    precio_rango = request.GET.get('precio', '')
    
    productoListado = Producto.objects.all()
    
    if nombre:
        productoListado = productoListado.filter(nombreP__icontains=nombre)
    
    if categoria_id:
        productoListado = productoListado.filter(categoria_id=categoria_id)
    
    if precio_rango:
        min_precio, max_precio = map(int, precio_rango.split('-'))
        productoListado = productoListado.filter(precio__gte=min_precio, precio__lte=max_precio)
    
    categorias = Categoria.objects.all()
    carrito = ItemCarrito.objects.filter(usuario=request.user)
    total = sum(item.producto.precio * item.cantidad for item in carrito)
    carrito_count = sum(item.cantidad for item in carrito)
    contexto = {
        "productos": productoListado,
        "categorias": categorias,
        'carrito': carrito,
        'total': total,
        'carrito_count': carrito_count,
    }    
    return render(request, 'core/productos.html', contexto)

def quienessomos(request):
    return render(request, 'core/quienessomos.html')

def galeria(request):
    return render(request, 'core/galeria.html')

def detalleproducto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    carrito = ItemCarrito.objects.filter(usuario=request.user)
    total = sum(item.producto.precio * item.cantidad for item in carrito)
    carrito_count = sum(item.cantidad for item in carrito)

    context = {
        'producto': producto,
        'carrito': carrito,
        'total': total,
        'carrito_count': carrito_count,
    }

    return render(request, 'core/detalleproducto.html', context)


@login_required
def agregar_al_carrito(request, producto_cod):
    producto = get_object_or_404(Producto, codProducto=producto_cod)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 1))  # Obtener la cantidad del formulario, por defecto 1 si no se proporciona

        # Obtener o crear el ítem en el carrito
        carrito, created = ItemCarrito.objects.get_or_create(usuario=request.user, producto=producto)

        if created:
            carrito.cantidad = cantidad  # Establecer la cantidad seleccionada
        else:
            carrito.cantidad += cantidad  # Sumar la cantidad seleccionada

        carrito.save()
        messages.success(request, 'Producto añadido al carrito correctamente.')

    return redirect('detalleproducto', pk=producto.pk)

@login_required
def aumentar_cantidad(request, producto_cod):
    producto = get_object_or_404(Producto, codProducto=producto_cod)
    
    if request.method == 'POST':
        carrito, created = ItemCarrito.objects.get_or_create(usuario=request.user, producto=producto)
        carrito.cantidad += 1
        carrito.save()
        messages.success(request, 'Cantidad aumentada correctamente.')
    
    # Redireccionar a la misma página o a donde corresponda después del aumento
    return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

@login_required
def disminuir_cantidad(request, producto_cod):
    producto = get_object_or_404(Producto, codProducto=producto_cod)
    
    if request.method == 'POST':
        carrito = get_object_or_404(ItemCarrito, usuario=request.user, producto=producto)

        if carrito.cantidad > 1:
            carrito.cantidad -= 1
            carrito.save()
            messages.success(request, 'Cantidad disminuida correctamente.')
        else:
            messages.warning(request, 'La cantidad no puede ser menor que 1. (PUEDES ELIMINAR EL PRODUCTO DEL CARRITO)')
    
    # Redireccionar a la misma página o a donde corresponda después de la disminución
    return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

@login_required
def carrito(request):
    carrito = ItemCarrito.objects.filter(usuario=request.user)
    total = sum(item.producto.precio * item.cantidad for item in carrito)
    carrito_count = sum(item.cantidad for item in carrito)

    regiones = Region.objects.all()

    # Obtener la región seleccionada, si existe
    region_id = request.GET.get('region')
    comunas = []
    if region_id:
        try:
            region_seleccionada = Region.objects.get(idRegion=region_id)
            comunas = Comuna.objects.filter(region=region_seleccionada)
        except Region.DoesNotExist:
            pass

    context = {
        'carrito': carrito,
        'total': total,
        'carrito_count': carrito_count,
        'regiones': regiones,
        'comunas': comunas,
        'region_id': region_id,
    }
    return render(request, 'core/carrito.html', context)


def comunas_por_region(request, region_id):
    try:
        region = Region.objects.get(idRegion=region_id)  # Ajuste aquí
        comunas = Comuna.objects.filter(region=region)
        data = [{'id': comuna.idComuna, 'nombreC': comuna.nombreC} for comuna in comunas]  # Ajuste aquí
        return JsonResponse(data, safe=False)
    except Region.DoesNotExist:
        return JsonResponse([], safe=False)


@login_required
def eliminar_del_carrito(request, carrito_id):
    item = get_object_or_404(ItemCarrito, pk=carrito_id)
    item.delete()
    messages.success(request, 'Producto eliminado correctamente.')
    return redirect('carrito')

@login_required
def crear_venta(request):
    if request.method == 'POST':
        tipo_entrega = request.POST.get('tipo_entrega')
        user = request.user
        usuario = Usuario.objects.get(correo=user.username)
        total = request.POST.get('total')

        if tipo_entrega == 'tienda':

            calle = 'Zenteno'
            numero = 524
            comuna = Comuna.objects.get(idComuna= 2)

            nueva_direccion = Direccion(calle=calle, numero=numero, comuna=comuna)
            estado = Estado.objects.get(id_estado = 1)
            nueva_direccion.save()
            tipo_despacho = TipoDespacho.objects.get(nombreDespacho='Tienda')
            nueva_venta = Venta(usuario=usuario, estadoP=estado, tipodespacho=tipo_despacho, total=total, direccion=nueva_direccion)
            
            nueva_venta.save()
            messages.success(request, 'Compra realizada correctamente!.')

            # Crear detalles de venta para cada producto en el carrito
            carrito_items = ItemCarrito.objects.filter(usuario=request.user)
            for item in carrito_items:
                detalle_venta = DetalleVenta(venta=nueva_venta, producto=item.producto, cantidad=item.cantidad, subtotal=item.producto.precio * item.cantidad)
                detalle_venta.save()

            # Limpiar el carrito después de completar la venta
            carrito_items.delete()

        elif tipo_entrega == 'domicilio':
            comuna_id = request.POST.get('comuna')
            calle = request.POST.get('calle')
            numero = request.POST.get('numero')
            comuna = Comuna.objects.get(idComuna=comuna_id)

            # Crear la dirección para la venta
            nueva_direccion = Direccion(calle=calle, numero=numero, comuna=comuna)
            nueva_direccion.save()

            # Obtener el tipo de despacho (a domicilio)
            tipo_despacho = TipoDespacho.objects.get(nombreDespacho='Domicilio')

            # Crear la venta con la dirección asociada
            estado = Estado.objects.get(id_estado = 1)
            nueva_venta = Venta(usuario=usuario, estadoP=estado, tipodespacho=tipo_despacho, total=total, direccion=nueva_direccion)
            nueva_venta.save()
            messages.success(request, 'Compra realizada correctamente!.')
            


            # Crear detalles de venta para cada producto en el carrito
            carrito_items = ItemCarrito.objects.filter(usuario=request.user)
            for item in carrito_items:
                detalle_venta = DetalleVenta(venta=nueva_venta, producto=item.producto, cantidad=item.cantidad, subtotal=item.producto.precio * item.cantidad)
                detalle_venta.save()

            # Limpiar el carrito después de completar la venta
            carrito_items.delete()

        return redirect('carrito')

    return redirect('carrito')

def listaventas(request):
    ventas_listado = Venta.objects.all()
    contexto = {
        "ventas": ventas_listado
    }
    return render(request, 'core/listaventas.html', contexto)

def detalles_venta(request):
    if request.method == 'GET' and 'venta_id' in request.GET:
        venta_id = request.GET.get('venta_id')
        detalles_venta = DetalleVenta.objects.filter(venta_id=venta_id)
        detalles = []
        for detalle in detalles_venta:
            producto = detalle.producto
            detalles.append({
                'producto': detalle.producto.nombreP,
                'imagen': producto.foto.url if producto.foto else '', 
                'cantidad': detalle.cantidad,
                'precio_unitario': detalle.producto.precio,
                'subtotal': detalle.subtotal
            })
        return JsonResponse({'detalles_venta': detalles})
    else:
        return JsonResponse({'error': 'No se encontró la venta solicitada'}, status=400)

def cambiar_estado_venta(request):
    if request.method == 'POST':
        venta_id = request.POST.get('venta_id')
        try:
            venta = Venta.objects.get(pk=venta_id)
            estado_actual = venta.estadoP
            if estado_actual.id_estado == 1: 
                nuevo_estado = Estado.objects.get(id_estado=2)
                clase_boton = 'btn-secondary'
            elif estado_actual.id_estado == 2:
                nuevo_estado = Estado.objects.get(id_estado=3)  
                clase_boton = 'btn-primary'
            elif estado_actual.id_estado == 3:  
                nuevo_estado = Estado.objects.get(id_estado=4)
                clase_boton = 'btn-success'
            else:
                nuevo_estado = Estado.objects.get(id_estado=1)
                clase_boton = 'btn-warning'

            venta.estadoP = nuevo_estado
            venta.save()

            return JsonResponse({
                'estado': {
                    'id_estado': nuevo_estado.id_estado,
                    'nombreEs': nuevo_estado.nombreEs,
                    'claseBoton': clase_boton
                }
            })
        except Venta.DoesNotExist:
            return JsonResponse({'error': 'Venta no encontrada'}, status=404)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def add_review(request):
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.user = request.user
            review.save()
            return redirect('inicio')
    else:
        form = ReviewForm()
    return render(request, 'core/add_review.html', {'form': form})


